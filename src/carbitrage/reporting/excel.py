"""Write a comparison to a spreadsheet.

Cases are defined in Python; this module is the way out, not the way in.  It
exists because the people who have to act on a capital budgeting result mostly
live in spreadsheets, and handing them a workbook they can pivot and re-format
is more useful than handing them a terminal table.

What it writes is a *report*, not a model: the numbers are values, not formulas,
because a formula written here would be a second implementation of the engine
and the two would drift apart.  The cash-flow sheet carries the full period grid
so the arithmetic can be checked by hand.

Requires the ``excel`` extra::

    pip install "carbitrage[excel]"
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

from ..cashflow import Component

if TYPE_CHECKING:  # pragma: no cover
    from ..comparison import ComparisonResult
    from ..scenario import ScenarioAnalysis

__all__ = ["write_excel"]

_MONEY = "#,##0.00"
_RATE = "0.0000"
_HEADER_FILL = "DDEBF7"


def write_excel(
    result: ComparisonResult,
    path: str | Path,
    *,
    baseline: str | None = None,
    scenarios: ScenarioAnalysis | None = None,
    include_cashflows: bool = True,
) -> Path:
    """Write ``result`` to an ``.xlsx`` workbook and return the path.

    Sheets written:

    * **Result** — the ranked table: present value, equivalent annual cost, cost
      per kilometre and the delta against the baseline.
    * **Breakdown** — every labelled component of every alternative, with a
      checked total per alternative.
    * **Cashflow** — the full period grid, one column per alternative, plus the
      discount factor, so the present values can be recomputed by hand.
    * **Scenarios** — present value by scenario with expected value and maximum
      regret, when a :class:`~carbitrage.study.scenario.ScenarioAnalysis` is supplied.

    Args:
        result: The comparison to report.
        path: Destination file.  A ``.xlsx`` suffix is added when missing.
        baseline: Alternative the deltas are measured against.
        scenarios: Optional scenario analysis to append.
        include_cashflows: Write the period-by-period grid.

    Raises:
        ImportError: if the ``excel`` extra is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on the environment
        raise ImportError(
            "write_excel() needs openpyxl; install carbitrage with the 'excel' extra"
        ) from exc

    destination = Path(path)
    if destination.suffix.lower() != ".xlsx":
        destination = destination.with_suffix(".xlsx")

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    wrap = Alignment(wrap_text=True, vertical="top")

    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]

    def header(sheet: Any, row: int, labels: list[str]) -> None:
        for column, label in enumerate(labels, start=1):
            cell = sheet.cell(row=row, column=column, value=label)
            cell.font = bold
            cell.fill = fill
            cell.alignment = wrap

    def autosize(sheet: Any, widths: dict[int, int]) -> None:
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = width

    # ------------------------------------------------------------- Result
    sheet = workbook.create_sheet("Result")
    sheet["A1"] = "Capital budgeting comparison"
    sheet["A1"].font = Font(bold=True, size=13)
    sheet["A2"] = (
        "Decision rule: the highest net present value, equivalently the lowest present "
        "value of outflows.  Outflows are negative in the cash-flow layer; the "
        "'PV of outflows' column flips the sign for reading only."
    )
    sheet["A2"].alignment = wrap
    sheet["A4"] = result.verdict()
    sheet["A4"].font = bold

    rows = result.to_rows(baseline)
    base_name = baseline if baseline is not None else result.ranking()[-1].name
    labels = [
        "Rank",
        "Alternative",
        "NPV",
        "PV of outflows",
        "EAC per year",
        "Cost per km",
        f"Delta vs. {base_name}",
    ]
    header(sheet, 6, labels)
    for index, row in enumerate(rows, start=7):
        sheet.cell(row=index, column=1, value=row["rank"])
        sheet.cell(row=index, column=2, value=row["alternative"])
        for column, key, fmt in (
            (3, "npv", _MONEY),
            (4, "pv_of_outflows", _MONEY),
            (5, "eac", _MONEY),
            (6, "cost_per_km", _RATE),
            (7, "delta_vs_baseline", _MONEY),
        ):
            value = row[key]
            if value is None:
                continue
            cell = sheet.cell(row=index, column=column, value=float(value))
            cell.number_format = fmt
    autosize(sheet, {1: 6, 2: 30, 3: 16, 4: 16, 5: 15, 6: 12, 7: 20})

    constraints = result.constraints()
    if constraints:
        start = 8 + len(rows)
        sheet.cell(row=start, column=1, value="Constraints").font = bold
        offset = start + 1
        for name, items in constraints.items():
            for item in items:
                sheet.cell(row=offset, column=1, value=name)
                sheet.cell(row=offset, column=2, value=item).alignment = wrap
                offset += 1

    # ---------------------------------------------------------- Breakdown
    sheet = workbook.create_sheet("Breakdown")
    sheet["A1"] = "Where each present value comes from"
    sheet["A1"].font = Font(bold=True, size=13)
    sheet["A2"] = "Each block sums to the alternative's net present value."
    header(sheet, 4, ["Alternative", "Component", "Line", "PV"])
    row_index = 5
    for name in result.names:
        total = 0.0
        for line in result.detail(name):
            if line.pv == 0.0:
                continue
            sheet.cell(row=row_index, column=1, value=name)
            sheet.cell(row=row_index, column=2, value=str(Component(line.label).value))
            sheet.cell(row=row_index, column=3, value=line.description)
            cell = sheet.cell(row=row_index, column=4, value=line.pv)
            cell.number_format = _MONEY
            total += line.pv
            row_index += 1
        sheet.cell(row=row_index, column=3, value="Total").font = bold
        cell = sheet.cell(row=row_index, column=4, value=total)
        cell.number_format = _MONEY
        cell.font = bold
        row_index += 2
    autosize(sheet, {1: 30, 2: 16, 3: 40, 4: 16})

    # ----------------------------------------------------------- Cashflow
    if include_cashflows:
        sheet = workbook.create_sheet("Cashflow")
        timeline = result.evaluations[0].ctx.timeline
        sheet["A1"] = "Period cash flows"
        sheet["A1"].font = Font(bold=True, size=13)
        sheet["A2"] = (
            "Negative is an outflow.  Acquisition and disposal at period 0, running costs "
            "in arrears, terminal values in the final period."
        )
        header(sheet, 4, ["Period", "Years", "Discount factor", *result.names])
        factors = timeline.discount_factors()
        columns = {name: result[name].amounts for name in result.names}
        for t in range(timeline.n_periods + 1):
            grid_row = 5 + t
            sheet.cell(row=grid_row, column=1, value=t)
            sheet.cell(row=grid_row, column=2, value=round(timeline.years_at(t), 4))
            cell = sheet.cell(row=grid_row, column=3, value=float(factors[t]))
            cell.number_format = "0.000000"
            for offset, name in enumerate(result.names, start=4):
                cell = sheet.cell(row=grid_row, column=offset, value=float(columns[name][t]))
                cell.number_format = _MONEY
        total_row = 6 + timeline.n_periods
        sheet.cell(row=total_row, column=1, value="NPV").font = bold
        for offset, name in enumerate(result.names, start=4):
            cell = sheet.cell(row=total_row, column=offset, value=result[name].npv)
            cell.number_format = _MONEY
            cell.font = bold
        autosize(sheet, {1: 8, 2: 8, 3: 15, **dict.fromkeys(range(4, 4 + len(result.names)), 20)})

    # ---------------------------------------------------------- Scenarios
    if scenarios is not None:
        sheet = workbook.create_sheet("Scenarios")
        sheet["A1"] = "Present value by scenario"
        sheet["A1"].font = Font(bold=True, size=13)
        scenario_names = [s.name for s in scenarios.scenarios]
        header(sheet, 3, ["Alternative", *scenario_names, "Expected", "Max regret"])
        table = scenarios.table()
        expected = scenarios.expected_npv()
        worst = scenarios.max_regret()
        for index, name in enumerate(scenarios.names, start=4):
            sheet.cell(row=index, column=1, value=name)
            for column, scenario_name in enumerate(scenario_names, start=2):
                cell = sheet.cell(row=index, column=column, value=table[name][scenario_name])
                cell.number_format = _MONEY
            cell = sheet.cell(row=index, column=2 + len(scenario_names), value=expected[name])
            cell.number_format = _MONEY
            cell = sheet.cell(row=index, column=3 + len(scenario_names), value=worst[name])
            cell.number_format = _MONEY
        footer = 5 + len(scenarios.names)
        sheet.cell(row=footer, column=1, value="Highest expected value").font = bold
        sheet.cell(row=footer, column=2, value=scenarios.best_by_expected_value())
        sheet.cell(row=footer + 1, column=1, value="Lowest worst-case regret").font = bold
        sheet.cell(row=footer + 1, column=2, value=scenarios.best_by_minimax_regret())
        autosize(sheet, {1: 30, **dict.fromkeys(range(2, 4 + len(scenario_names)), 18)})

    workbook.save(destination)
    return destination
